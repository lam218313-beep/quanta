import sys
import os
import subprocess
from pathlib import Path
from fastapi import FastAPI, BackgroundTasks, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import asyncio
import threading
import traceback
import io
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI()

# Allow CORS so React (port 5173) can talk to this API (port 8000)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Allow all for local dev
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class BotRequest(BaseModel):
    ruc: str
    periodo: str | None = None
    tipo_libro: str | None = None

# Store references to running tasks to prevent overlapping
running_tasks = {}

# Ensure logs directory exists
LOGS_DIR = Path(__file__).parent / "logs"
LOGS_DIR.mkdir(exist_ok=True)

def _run_sync_process(task_id: str, command: list, cwd: str):
    """Runs a command capturing stdout via PIPE (for headless/CLI processes)."""
    log_file = LOGS_DIR / f"{task_id}.log"
    with open(log_file, "w", encoding="utf-8") as f:
        f.write(f"[{task_id}] Starting command: {' '.join(command)}\n")
        print(f"[{task_id}] Starting command: {' '.join(command)}")
        
        try:
            process = subprocess.Popen(
                command,
                cwd=cwd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding='utf-8',
                errors='replace'
            )
            # Store the actual process object so it can be killed
            if task_id in running_tasks:
                running_tasks[task_id] = process
                
            for line in process.stdout:
                f.write(line)
                f.flush()
                print(f"[{task_id}] {line.strip()}")
                
            process.wait()
            f.write(f"[{task_id}] Finished with return code {process.returncode}\n")
            print(f"[{task_id}] Finished with return code {process.returncode}")
        except Exception as e:
            err_trace = traceback.format_exc()
            f.write(f"[{task_id}] Failed to run command: {e}\n{err_trace}\n")
            print(f"[{task_id}] Failed to run command: {e}")
            print(err_trace)
        finally:
            if task_id in running_tasks:
                del running_tasks[task_id]

def _run_headed_process(task_id: str, command: list, cwd: str):
    """
    Runs a GUI/headed process (e.g. Playwright with headless=False) writing stdout
    DIRECTLY to the log file instead of via subprocess.PIPE.
    This avoids the EPIPE 'broken pipe' crash that occurs when Playwright's internal
    Node.js driver writes events back through a captured pipe.
    """
    log_file = LOGS_DIR / f"{task_id}.log"
    with open(log_file, "w", encoding="utf-8") as log_fh:
        log_fh.write(f"[{task_id}] Starting headed command: {' '.join(command)}\n")
        log_fh.flush()
        print(f"[{task_id}] Starting headed command: {' '.join(command)}")
        
        try:
            # Write directly to file — no PIPE, no EPIPE
            process = subprocess.Popen(
                command,
                cwd=cwd,
                stdout=log_fh,
                stderr=log_fh,
                text=True,
                encoding='utf-8',
                errors='replace',
                env={**os.environ, "PYTHONUNBUFFERED": "1"}  # force line-buffered output
            )
            if task_id in running_tasks:
                running_tasks[task_id] = process
                
            process.wait()
            log_fh.write(f"[{task_id}] Finished with return code {process.returncode}\n")
            log_fh.flush()
            print(f"[{task_id}] Headed process finished with return code {process.returncode}")
        except Exception as e:
            err_trace = traceback.format_exc()
            log_fh.write(f"[{task_id}] Failed to run headed command: {e}\n{err_trace}\n")
            log_fh.flush()
            print(f"[{task_id}] Failed to run headed command: {e}")
        finally:
            if task_id in running_tasks:
                del running_tasks[task_id]

async def run_command_in_background(task_id: str, command: list, cwd: str, headed: bool = False):
    target = _run_headed_process if headed else _run_sync_process
    thread = threading.Thread(target=target, args=(task_id, command, cwd))
    thread.daemon = True
    thread.start()

@app.get("/api/bot/status")
def get_running_tasks():
    """Returns all currently running task IDs."""
    tasks = []
    for task_id, proc in running_tasks.items():
        pid = proc.pid if hasattr(proc, 'pid') else None
        tasks.append({"task_id": task_id, "pid": pid})
    return {"running_tasks": tasks, "count": len(tasks)}

@app.post("/api/bot/reset")
def reset_running_tasks(task_id: str = None):
    """
    Kills and clears stuck running tasks.
    If task_id is provided, only clears that task.
    Otherwise clears ALL running tasks.
    """
    killed = []
    to_remove = [task_id] if task_id and task_id in running_tasks else list(running_tasks.keys())
    
    for tid in to_remove:
        proc = running_tasks.get(tid)
        if proc and hasattr(proc, 'kill'):
            try:
                proc.kill()
                killed.append(tid)
            except Exception:
                pass
        if tid in running_tasks:
            del running_tasks[tid]
            if tid not in killed:
                killed.append(tid)
                
    return {"cleared": killed, "message": f"Cleared {len(killed)} stuck task(s). You can now restart the bot."}

@app.post("/api/bot/download-api")
async def trigger_download_api(req: BotRequest, background_tasks: BackgroundTasks):
    if not req.ruc or not req.periodo:
        raise HTTPException(status_code=400, detail="RUC and Periodo are required")
        
    task_id = f"api_{req.ruc}_{req.periodo}"
    if task_id in running_tasks:
        return {"status": "already_running", "message": "This API task is already running.", "task_id": task_id}
        
    cmd = [sys.executable, "app/brain/sire_download_cli.py", "--client", req.ruc, "--period", req.periodo]
    root_dir = Path(__file__).parent.parent
    
    running_tasks[task_id] = True
    background_tasks.add_task(run_command_in_background, task_id, cmd, str(root_dir))
    return {"status": "started", "message": f"Started SIRE API Download for {req.ruc} - {req.periodo}", "task_id": task_id}


# ─────────────────────────────────────────────
# DESCARGA MASIVA DE PRELIMINAR SIRE (BATCH)
# ─────────────────────────────────────────────

class BatchDownloadRequest(BaseModel):
    periodo: str
    rucs: list[str] | None = None  # Si None → usa todos los clientes con credenciales


def _run_batch_download(task_id: str, rucs: list[str], periodo: str):
    """
    Worker secuencial: descarga el preliminar SIRE de cada RUC en orden,
    uno por uno. Escribe logs acumulativos en un único archivo .log.
    """
    log_file = LOGS_DIR / f"{task_id}.log"
    root_dir = Path(__file__).parent.parent
    total = len(rucs)

    with open(log_file, "w", encoding="utf-8") as f:
        f.write(f"{'='*60}\n")
        f.write(f"  INICIO DESCARGA MASIVA DE PRELIMINAR SIRE\n")
        f.write(f"  Periodo: {periodo}  |  Total clientes: {total}\n")
        f.write(f"{'='*60}\n\n")
        f.flush()

        exitosos = 0
        fallidos = 0

        for i, ruc in enumerate(rucs, 1):
            f.write(f"\n{'─'*50}\n")
            f.write(f"[PROGRESO] {i}/{total} ── Cliente RUC: {ruc}\n")
            f.write(f"{'─'*50}\n")
            f.flush()

            cmd = [
                sys.executable,
                "app/brain/sire_download_cli.py",
                "--client", ruc,
                "--period", periodo,
            ]

            try:
                process = subprocess.Popen(
                    cmd,
                    cwd=str(root_dir),
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    env={**os.environ, "PYTHONUNBUFFERED": "1"},
                )

                for line in process.stdout:
                    f.write(line)
                    f.flush()

                process.wait()
                rc = process.returncode

                if rc == 0:
                    exitosos += 1
                    f.write(f"[OK] Cliente {ruc} finalizado correctamente.\n")
                else:
                    fallidos += 1
                    f.write(f"[ERROR] Cliente {ruc} terminó con código {rc}.\n")

            except Exception as e:
                fallidos += 1
                f.write(f"[EXCEPCION] Cliente {ruc}: {e}\n")

            f.flush()
            
            # Descanso de 15 segundos entre clientes para evitar 429 Too Many Requests de SUNAT
            if i < total:
                f.write(f"Esperando 15 segundos para no saturar a SUNAT...\n")
                f.flush()
                time.sleep(15)

        f.write(f"\n{'='*60}\n")
        f.write(f"  FIN DESCARGA MASIVA\n")
        f.write(f"  Exitosos: {exitosos}/{total}  |  Fallidos: {fallidos}/{total}\n")
        f.write(f"{'='*60}\n")
        f.flush()

    if task_id in running_tasks:
        del running_tasks[task_id]


@app.post("/api/bot/batch-download-api")
async def trigger_batch_download_api(req: BatchDownloadRequest, background_tasks: BackgroundTasks):
    """
    Descarga el preliminar SIRE para múltiples clientes en cola, uno por uno.
    Si 'rucs' es None, descarga para todos los clientes que tienen usuario_sol configurado.
    """
    if not req.periodo:
        raise HTTPException(status_code=400, detail="Periodo es requerido")

    task_id = f"batch_api_{req.periodo}"
    if task_id in running_tasks:
        return {
            "status": "already_running",
            "message": "Ya hay una descarga masiva en curso para este periodo.",
            "task_id": task_id,
        }

    # Determinar la lista de RUCs a procesar
    rucs_to_process = req.rucs
    if not rucs_to_process:
        # Obtener todos los clientes con credenciales SOL y API configuradas
        from app.brain.db.supabase_client import get_supabase
        sb = get_supabase()
        res = sb.table("clientes").select("ruc, usuario_sol, clave_sol, client_id_api, client_secret_api").execute()
        clientes_data = res.data or []
        rucs_to_process = [
            c["ruc"]
            for c in clientes_data
            if c.get("usuario_sol") and c.get("clave_sol") and c.get("client_id_api") and c.get("client_secret_api")
        ]

    if not rucs_to_process:
        raise HTTPException(
            status_code=400,
            detail="No hay clientes con credenciales SOL configuradas para procesar.",
        )

    running_tasks[task_id] = True
    background_tasks.add_task(_run_batch_download, task_id, rucs_to_process, req.periodo)

    return {
        "status": "started",
        "message": f"Descarga masiva iniciada: {len(rucs_to_process)} clientes para el periodo {req.periodo}.",
        "task_id": task_id,
        "total_clientes": len(rucs_to_process),
    }


@app.get("/api/clientes/con-credenciales")
def get_clientes_con_credenciales():
    """Devuelve la lista de clientes que tienen usuario_sol, clave_sol y credenciales API configurados."""
    from app.brain.db.supabase_client import get_supabase
    sb = get_supabase()
    res = sb.table("clientes").select("id, ruc, razon_social, usuario_sol, client_id_api").execute()
    clientes_data = res.data or []
    filtrados = [
        {"id": c["id"], "ruc": c["ruc"], "razon_social": c["razon_social"]}
        for c in clientes_data
        if c.get("usuario_sol") and c.get("client_id_api")
    ]
    return {"clientes": filtrados, "total": len(filtrados)}

@app.post("/api/bot/automation-login")
async def trigger_automation_login(req: BotRequest, background_tasks: BackgroundTasks):
    if not req.ruc:
        raise HTTPException(status_code=400, detail="RUC is required")
        
    task_id = f"login_{req.ruc}"
    if task_id in running_tasks:
        return {"status": "already_running", "message": "Login task is already running.", "task_id": task_id}
        
    cmd = [sys.executable, "-u", "app/brain/automation_scraper.py", "--ruc", req.ruc]
    root_dir = Path(__file__).parent.parent
    
    running_tasks[task_id] = True
    # Use headed=True to avoid EPIPE crash with Playwright's headed browser
    background_tasks.add_task(run_command_in_background, task_id, cmd, str(root_dir), True)
    return {"status": "started", "message": f"Started Authentication bot for {req.ruc}", "task_id": task_id}

@app.post("/api/bot/download-fisicos")
async def trigger_download_fisicos(req: BotRequest, background_tasks: BackgroundTasks):
    if not req.ruc:
        raise HTTPException(status_code=400, detail="RUC is required")
        
    task_id = f"fisicos_{req.ruc}"
    if req.tipo_libro:
        task_id += f"_{req.tipo_libro}"
    
    if task_id in running_tasks:
        return {"status": "already_running", "message": "XML Scraper is already running.", "task_id": task_id}
        
    cmd = [sys.executable, "app/brain/db/sire_bot_orchestrator.py", "--limit", "200", "--ruc", req.ruc]
    if req.periodo:
        cmd += ["--periodo", req.periodo]
    if req.tipo_libro:
        cmd += ["--tipo_libro", req.tipo_libro]
    root_dir = Path(__file__).parent.parent
    
    running_tasks[task_id] = True
    background_tasks.add_task(run_command_in_background, task_id, cmd, str(root_dir))
    
    msg_suffix = f" - {req.periodo or 'todos los periodos'}"
    if req.tipo_libro:
        msg_suffix += f" (Solo {req.tipo_libro})"
        
    return {"status": "started", "message": f"Started XML Download bot for {req.ruc}{msg_suffix}", "task_id": task_id}

@app.post("/api/bot/sync-files")
async def trigger_sync_files(background_tasks: BackgroundTasks):
    """Scans all local download folders and reconciles with Supabase DB."""
    task_id = "sync_files"
    if task_id in running_tasks:
        return {"status": "already_running", "message": "Sync is already running.", "task_id": task_id}
        
    cmd = [sys.executable, "app/brain/db/sync_files.py"]
    root_dir = Path(__file__).parent.parent
    
    running_tasks[task_id] = True
    background_tasks.add_task(run_command_in_background, task_id, cmd, str(root_dir))
    return {"status": "started", "message": "Sincronizando archivos físicos con base de datos...", "task_id": task_id}

@app.get("/api/bot/local-files")
def get_local_files():
    """Devuelve una lista plana de todos los nombres de archivo (.xml, .pdf, .zip) en la carpeta local downloads."""
    root_dir = Path(__file__).parent.parent
    downloads_dir = root_dir / "downloads"
    
    if not downloads_dir.exists():
        return {"files": []}
        
    files = []
    try:
        # Escaneo ultra rápido de la carpeta de descargas
        for f in downloads_dir.rglob("*.*"):
            if f.suffix.lower() in ('.xml', '.pdf', '.zip'):
                files.append(f.name)
    except Exception:
        pass
        
    return {"files": list(set(files))}

@app.post("/api/bot/upload-manual")
async def upload_manual_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    ruc_tercero: str = Form(...),
    tipo_cp: str = Form(...),
    serie: str = Form(...),
    numero: str = Form(...),
    cliente_ruc: str = Form(...)
):
    import shutil
    try:
        root_dir = Path(__file__).parent.parent
        downloads_dir = root_dir / "downloads" / "manual_uploads"
        downloads_dir.mkdir(parents=True, exist_ok=True)
        
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ['.xml', '.zip', '.pdf']:
            raise HTTPException(status_code=400, detail="El archivo debe ser .xml, .zip o .pdf")
            
        ruc_terc = ruc_tercero.strip()
        if not ruc_terc or ruc_terc == '-':
            ruc_terc = cliente_ruc
            
        filename = f"{ruc_terc}-{tipo_cp}-{serie}-{numero}{ext}"
        filepath = downloads_dir / filename
        
        with open(filepath, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # Ejecutar sync_files para actualizar la base de datos
        cmd = [sys.executable, "app/brain/db/sync_files.py"]
        background_tasks.add_task(run_command_in_background, f"sync_manual_{time.time()}", cmd, str(root_dir))
        
        return {"status": "ok", "message": "Archivo subido y sincronización iniciada", "filename": filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/bot/enrich-xml")
async def trigger_enrich_xml(req: BotRequest, background_tasks: BackgroundTasks):
    if not req.ruc:
        raise HTTPException(status_code=400, detail="RUC is required")
        
    task_id = f"enrich_{req.ruc}"
    if task_id in running_tasks:
        return {"status": "already_running", "message": "XML Enricher is already running.", "task_id": task_id}
        
    cmd = [sys.executable, "app/brain/db/sire_xml_enricher.py", "--limit", "500", "--ruc", req.ruc]
    if req.periodo:
        cmd += ["--periodo", req.periodo]
    root_dir = Path(__file__).parent.parent
    
    running_tasks[task_id] = True
    background_tasks.add_task(run_command_in_background, task_id, cmd, str(root_dir))
    return {"status": "started", "message": f"Started XML Extraction bot for {req.ruc} - {req.periodo or 'todos los periodos'}", "task_id": task_id}

@app.post("/api/bot/classify-ai")
async def trigger_classify_ai(req: BotRequest, background_tasks: BackgroundTasks):
    if not req.ruc:
        raise HTTPException(status_code=400, detail="RUC is required")
        
    task_id = f"classify_{req.ruc}"
    if task_id in running_tasks:
        return {"status": "already_running", "message": "AI Classifier is already running.", "task_id": task_id}
        
    cmd = [sys.executable, "app/brain/db/ai_classifier.py", "--limit", "100", "--ruc", req.ruc]
    if req.periodo:
        cmd += ["--periodo", req.periodo]
    root_dir = Path(__file__).parent.parent
    
    running_tasks[task_id] = True
    background_tasks.add_task(run_command_in_background, task_id, cmd, str(root_dir))
    return {"status": "started", "message": f"Started AI Classifier bot for {req.ruc} - {req.periodo or 'todos los periodos'}", "task_id": task_id}

@app.get("/api/bot/logs/{task_id}")
def get_task_logs(task_id: str):
    log_file = LOGS_DIR / f"{task_id}.log"
    is_running = task_id in running_tasks
    if not log_file.exists():
        return {"task_id": task_id, "logs": "No logs available yet...", "is_running": is_running}
    
    with open(log_file, "r", encoding="utf-8", errors="replace") as f:
        # For huge files, we should probably read the last N lines, but for these bots it should be fine.
        content = f.read()

    return {"task_id": task_id, "logs": content, "is_running": is_running}
    
import os
import tempfile
import fitz  # PyMuPDF
from fastapi.responses import FileResponse
from app.brain.db.supabase_client import get_supabase

class ExportPdfRequest(BaseModel):
    ruc: str
    periodo: str
    tipo_libro: str
    allow_incomplete: bool = False

def _safe_remove_file(path: str) -> None:
    try:
        os.remove(path)
    except Exception:
        # Best-effort cleanup; avoid surfacing Windows file-lock errors.
        pass

@app.post("/api/export/pdf-merged")
def export_pdf_merged(req: ExportPdfRequest, background_tasks: BackgroundTasks):
    supabase = get_supabase()
    
    # Obtener el id del cliente
    res_cli = supabase.table("clientes").select("id").eq("ruc", req.ruc).execute()
    if not res_cli.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente_id = res_cli.data[0]["id"]
    
    # Obtener las rutas de los pdf y datos del comprobante de ese cliente, periodo y tipo
    res_docs = supabase.table("sire_comprobantes_fisicos").select("id, serie, numero, ruta_pdf").eq("cliente_id", cliente_id).eq("periodo", req.periodo).eq("tipo_libro", req.tipo_libro).execute()
    
    if not res_docs.data:
        raise HTTPException(status_code=404, detail="No hay comprobantes para exportar")
        
    comprobantes = res_docs.data
    failed_comprobantes = []
    
    try:
        merged_doc = fitz.open()
        for comp in comprobantes:
            pdf_path = comp.get("ruta_pdf")
            serie = comp.get("serie", "N/A")
            numero = comp.get("numero", "N/A")
            
            if not pdf_path or not os.path.exists(pdf_path):
                failed_comprobantes.append(comp)
                continue
                
            try:
                doc = fitz.open(pdf_path)
                merged_doc.insert_pdf(doc)
                doc.close()
            except Exception as e:
                print(f"Error merging {pdf_path}: {e}")
                failed_comprobantes.append(comp)
                
        # Si todos fallaron, no hay nada que compilar
        if len(failed_comprobantes) == len(comprobantes):
            merged_doc.close()
            raise HTTPException(status_code=400, detail="Ningún comprobante tiene un PDF válido o descargado para compilar.")
            
        # Si hubo comprobantes que fallaron y NO se permite incompleto, abortar
        if failed_comprobantes and not req.allow_incomplete:
            merged_doc.close()
            failed_names = []
            
            for fcomp in failed_comprobantes:
                failed_names.append(f"{fcomp.get('serie')}-{fcomp.get('numero')}")
                # Actualizar a PENDIENTE para que el bot vuelva a intentarlo
                supabase.table("sire_comprobantes_fisicos").update({
                    "estado_xml": "PENDIENTE",
                    "estado_pdf": "PENDIENTE",
                    "reintentos": 0
                }).eq("id", fcomp["id"]).execute()
            
            error_msg = f"No se pudo generar el compilado PDF porque los siguientes comprobantes no se descargaron correctamente o están corruptos: {', '.join(failed_names)}. Se han devuelto a estado PENDIENTE."
            raise HTTPException(status_code=400, detail=error_msg)
            
        # Si se permite incompleto, solo marcar los fallidos como PENDIENTE sin abortar
        if failed_comprobantes and req.allow_incomplete:
            for fcomp in failed_comprobantes:
                supabase.table("sire_comprobantes_fisicos").update({
                    "estado_pdf": "PENDIENTE",
                    "reintentos": 0
                }).eq("id", fcomp["id"]).execute()

        # Guardar en un archivo temporal si todo está correcto
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        tmp.close()
        merged_doc.save(tmp.name)
        merged_doc.close()
        
        filename = f"Comprobantes_{req.tipo_libro}_{req.periodo}.pdf"
        background_tasks.add_task(_safe_remove_file, tmp.name)
        return FileResponse(tmp.name, filename=filename, media_type="application/pdf")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/bot/sync-files")
def sync_files_api():
    try:
        # Run sync_files.py synchronously and capture output
        cmd = [sys.executable, "app/brain/db/sync_files.py"]
        root_dir = Path(__file__).parent.parent
        result = subprocess.run(cmd, cwd=str(root_dir), capture_output=True, text=True, encoding='utf-8')
        return {"status": "ok", "message": "Archivos sincronizados", "output": result.stdout}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/health")
def health_check():
    return {"status": "ok", "running_tasks": list(running_tasks.keys())}




# ─────────────────────────────────────────────
# EXPORTAR PRELIMINAR SIRE (2 hojas: Ventas + Compras)
# ─────────────────────────────────────────────

class ExportPreliminarRequest(BaseModel):
    ruc: str
    periodo: str

class ExportSireTxtRequest(BaseModel):
    ruc: str
    periodo: str
    tipo_libro: str

@app.post("/api/export/sire-txt")
def export_sire_txt(req: ExportSireTxtRequest):
    """Genera el TXT personalizado para el sistema contable (M1)"""
    from app.brain.db.sire_txt_exporter import build_custom_compras_txt, build_custom_ventas_txt
    supabase = get_supabase()

    res_cli = supabase.table("clientes").select("id").eq("ruc", req.ruc).execute()
    if not res_cli.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente_id = res_cli.data[0]["id"]

    if req.tipo_libro == "COMPRAS":
        res = supabase.table("sire_preliminar_compras").select("*").eq("cliente_id", cliente_id).eq("periodo", req.periodo).order("fecha_emision").execute()
        txt_content = build_custom_compras_txt(res.data, req.ruc, req.periodo)
        filename = f"Compras_{req.periodo}_M1.txt"
    else:
        res = supabase.table("sire_preliminar_ventas").select("*").eq("cliente_id", cliente_id).eq("periodo", req.periodo).order("fecha_emision").execute()
        txt_content = build_custom_ventas_txt(res.data, req.ruc, req.periodo)
        filename = f"Ventas_{req.periodo}_M1.txt"

    from fastapi.responses import PlainTextResponse
    return PlainTextResponse(content=txt_content, media_type="text/plain", headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.post("/api/bot/upload-sire")
async def bot_upload_sire(req: ExportSireTxtRequest, background_tasks: BackgroundTasks):
    """
    Genera el TXT de reemplazo para SIRE, lo comprime en ZIP y llama al bot 
    para subirlo a SUNAT en segundo plano.
    """
    from app.brain.db.sire_txt_exporter import build_sire_compras_txt, build_sire_ventas_txt
    import zipfile
    import os
    import tempfile
    
    supabase = get_supabase()

    res_cli = supabase.table("clientes").select("*").eq("ruc", req.ruc).execute()
    if not res_cli.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente = res_cli.data[0]

    # 1. Generar TXT
    if req.tipo_libro == "COMPRAS":
        res = supabase.table("sire_preliminar_compras").select("*").eq("cliente_id", cliente["id"]).eq("periodo", req.periodo).order("fecha_emision").execute()
        txt_content = build_sire_compras_txt(res.data, req.ruc, req.periodo)
        filename_txt = f"LE{req.ruc}{req.periodo}000804000211112.txt"
    else:
        res = supabase.table("sire_preliminar_ventas").select("*").eq("cliente_id", cliente["id"]).eq("periodo", req.periodo).order("fecha_emision").execute()
        txt_content = build_sire_ventas_txt(res.data, req.ruc, req.periodo)
        filename_txt = f"LE{req.ruc}{req.periodo}001404000211112.txt"

    # 2. Guardar TXT y comprimir en ZIP
    tmp_dir = tempfile.gettempdir()
    zip_path = os.path.join(tmp_dir, filename_txt.replace(".txt", ".zip"))
    txt_path = os.path.join(tmp_dir, filename_txt)
    
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(txt_content)
        
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(txt_path, arcname=filename_txt)

    # 3. Llamar a la API de SIRE en background
    def run_sire_upload_api():
        from app.brain.integrations.sunat_sire import SunatSireClient
        import os
        
        print(f"Subiendo a SUNAT SIRE vía API para {req.ruc} - {req.periodo} ({req.tipo_libro})...")
        try:
            # Obtener credenciales (de BD o de entorno como fallback)
            client_id = cliente.get("api_client_id") or os.getenv("SUNAT_CLIENT_ID")
            client_secret = cliente.get("api_client_secret") or os.getenv("SUNAT_CLIENT_SECRET")
            ruc = cliente.get("ruc") or os.getenv("SUNAT_RUC")
            sol_user = cliente.get("sol_user") or os.getenv("SUNAT_USERNAME")
            sol_pass = cliente.get("sol_pass") or os.getenv("SUNAT_PASSWORD")

            if not client_id or not client_secret:
                print("Error: No hay credenciales API (client_id/client_secret) configuradas.")
                return

            client = SunatSireClient(
                client_id=client_id,
                client_secret=client_secret,
                ruc=ruc,
                username=sol_user,
                password=sol_pass
            )
            
            book_type = "sales" if req.tipo_libro == "VENTAS" else "purchases"
            
            ticket_id = client.upload_replacement(
                period=req.periodo,
                book_type=book_type,
                zip_path=zip_path
            )
            print(f"=== ÉXITO: PROPUESTA REEMPLAZADA ===")
            print(f"Ticket generado por SUNAT: {ticket_id}")
            
        except Exception as e:
            print(f"Error subiendo a la API de SUNAT: {e}")

    background_tasks.add_task(run_sire_upload_api)

    return {"ok": True, "message": "Bot de subida a SUNAT iniciado en segundo plano", "zip_path": zip_path}


@app.post("/api/export/preliminar-excel")
def export_preliminar_excel(req: ExportPreliminarRequest):
    """Genera un Excel con 2 hojas (VENTAS, COMPRAS) con los datos crudos del
    Preliminar SIRE. La hoja VENTAS incluye al final un cuadro de
    Liquidación de Impuestos (RENTA MYPE 1%, IGV 18%, Crédito Fiscal, etc.)
    """
    supabase = get_supabase()

    # ── Resolver cliente ────────────────────────────────────────────────────
    res_cli = supabase.table("clientes").select("id, razon_social").eq("ruc", req.ruc).execute()
    if not res_cli.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente_id = res_cli.data[0]["id"]
    razon_social = res_cli.data[0]["razon_social"]

    periodo_fmt = req.periodo  # ej: "202605"
    try:
        yr, mo = int(periodo_fmt[:4]), int(periodo_fmt[4:])
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        periodo_label = f"{meses[mo-1]} {yr}"
    except Exception:
        periodo_label = periodo_fmt

    # ── Obtener datos ───────────────────────────────────────────────────────
    res_v = supabase.table("sire_preliminar_ventas") \
        .select("*") \
        .eq("cliente_id", cliente_id) \
        .eq("periodo", req.periodo) \
        .order("fecha_emision") \
        .execute()

    res_c = supabase.table("sire_preliminar_compras") \
        .select("*") \
        .eq("cliente_id", cliente_id) \
        .eq("periodo", req.periodo) \
        .order("fecha_emision") \
        .execute()

    ventas  = res_v.data or []
    compras = res_c.data or []

    # ── Estilos de utilidad ─────────────────────────────────────────────────
    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _border_all():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def _bold_font(size=10, color="000000"):
        return Font(bold=True, size=size, color=color)

    HEADER_FILL  = _fill("1A3C5E")   # azul oscuro
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=9)
    ALT_ROW_FILL = _fill("EBF3FA")   # azul muy claro para filas pares

    # Colores del cuadro liquidación (tomados de la imagen)
    DARK_NAVY   = _fill("1A3C5E")
    LIGHT_GRAY  = _fill("D9D9D9")
    YELLOW_FILL = _fill("FFFF00")
    RED_FONT    = Font(bold=True, color="FF0000", size=10)
    WHITE_FONT  = Font(bold=True, color="FFFFFF", size=10)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT   = Font(bold=True, size=10)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PRELIMINAR_SIRE"

    # ════════════════════════════════════════════════════════════════════════
    # HEADER DATOS CLIENTE
    # ════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A1:K1")
    cell_title = ws.cell(row=1, column=1, value="CIERRE TRIBUTARIO MENSUAL")
    cell_title.font = _bold_font(size=14)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:E3")
    ws.cell(row=3, column=1, value=f"Razón Social: {razon_social}").font = _bold_font(size=11)
    ws.merge_cells("A4:E4")
    ws.cell(row=4, column=1, value=f"RUC: {req.ruc}").font = _bold_font(size=11)
    ws.merge_cells("A5:E5")
    ws.cell(row=5, column=1, value=f"Periodo: {periodo_label}").font = _bold_font(size=11)

    r = 7
    header_v_row = r

    # ════════════════════════════════════════════════════════════════════════
    # TABLA VENTAS
    # ════════════════════════════════════════════════════════════════════════
    headers_v = [
        "Fecha Emisión", "Serie", "Número",
        "RUC / DNI", "Razón Social",
        "Base Imponible", "IGV / IPM", "Exonerado", "Inafecto",
        "TOTAL", "Moneda"
    ]
    for c, h in enumerate(headers_v, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border_all()
    ws.row_dimensions[r].height = 28

    total_bi_v = 0.0
    total_igv_v = 0.0
    total_cp_v  = 0.0

    for i, row in enumerate(ventas, 1):
        r += 1
        fill = ALT_ROW_FILL if i % 2 == 0 else None
        vals = [
            row.get("fecha_emision") or "",
            row.get("serie_cdp") or "",
            row.get("nro_cp") or "",
            row.get("nro_doc_identidad") or "",
            row.get("razon_social") or "",
            float(row.get("bi_gravada") or 0),
            float(row.get("igv_ipm") or 0),
            float(row.get("mto_exonerado") or 0),
            float(row.get("mto_inafecto") or 0),
            float(row.get("total_cp") or 0),
            row.get("moneda") or "PEN",
        ]
        total_bi_v  += vals[5]
        total_igv_v += vals[6]
        total_cp_v  += vals[9]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = _border_all()
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            if c in (6, 7, 8, 9, 10):
                cell.number_format = '"S/"#,##0.00'

    # Fila de totales ventas
    total_row_v = r + 1
    r = total_row_v
    ws.cell(row=r, column=5, value="TOTAL VENTAS").font = _bold_font()
    ws.cell(row=r, column=5).fill = _fill("D9EDF7")
    for c, val in [(6, total_bi_v), (7, total_igv_v), (10, total_cp_v)]:
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = _bold_font()
        cell.fill = _fill("D9EDF7")
        cell.number_format = '"S/"#,##0.00'
        cell.border = _border_all()

    # ════════════════════════════════════════════════════════════════════════
    # TABLA COMPRAS
    # ════════════════════════════════════════════════════════════════════════
    r += 3
    r_compras_hdr = r

    headers_c = [
        "Fecha Emisión", "Serie", "Número",
        "RUC / DNI", "Razón Social",
        "BI Gravado", "IGV / IPM", "BI No Gravado", "Valor No Grav.",
        "TOTAL", "Moneda"
    ]
    for c, h in enumerate(headers_c, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = _fill("217346")
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border_all()
    ws.row_dimensions[r].height = 28

    total_bi_c = 0.0
    total_igv_c = 0.0
    total_cp_c2 = 0.0

    for i, row in enumerate(compras, 1):
        r += 1
        fill = ALT_ROW_FILL if i % 2 == 0 else None
        vals = [
            row.get("fecha_emision") or "",
            row.get("serie_cdp") or "",
            row.get("nro_cp") or "",
            row.get("nro_doc_identidad") or "",
            row.get("razon_social") or "",
            float(row.get("bi_gravado_dg") or 0),
            float(row.get("igv_ipm_dg") or 0),
            float(row.get("bi_gravado_dng") or 0),
            float(row.get("valor_adq_ng") or 0),
            float(row.get("total_cp") or 0),
            row.get("moneda") or "PEN",
        ]
        total_bi_c  += vals[5]
        total_igv_c += vals[6]
        total_cp_c2 += vals[9]

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = _border_all()
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            if c in (6, 7, 8, 9, 10):
                cell.number_format = '"S/"#,##0.00'

    # Fila totales compras
    total_row_c = r + 1
    r = total_row_c
    ws.cell(row=r, column=5, value="TOTAL COMPRAS").font = _bold_font()
    ws.cell(row=r, column=5).fill = _fill("C8E6C9")
    for c, val in [(6, total_bi_c), (7, total_igv_c), (10, total_cp_c2)]:
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = _bold_font()
        cell.fill = _fill("C8E6C9")
        cell.number_format = '"S/"#,##0.00'
        cell.border = _border_all()

    # ════════════════════════════════════════════════════════════════════════
    # CUADRO DE LIQUIDACIÓN DE IMPUESTOS
    # ════════════════════════════════════════════════════════════════════════
    def _liq_row(ws_target, row_idx, col_a, label, prefix, value, fill_a=None, fill_b=None,
                 font_label=None, font_val=None, num_fmt='"S/"#,##0.00'):
        c_lbl = ws_target.cell(row=row_idx, column=col_a,     value=label)
        c_pre = ws_target.cell(row=row_idx, column=col_a + 1, value=prefix)
        c_val = ws_target.cell(row=row_idx, column=col_a + 2, value=value)

        for c in (c_lbl, c_pre, c_val):
            c.border = _border_all()
        if fill_a:
            c_lbl.fill = fill_a
            c_pre.fill = fill_a
            c_val.fill = fill_a
        if fill_b:
            c_val.fill = fill_b
        if font_label:
            c_lbl.font = font_label
        if font_val:
            c_pre.font  = font_val
            c_val.font  = font_val
        c_val.number_format = num_fmt
        c_val.alignment = Alignment(horizontal="right")

    LIQ_START = r + 3
    LIQ_COL = 1
    r = LIQ_START

    # Título principal
    ws.merge_cells(start_row=r, start_column=LIQ_COL, end_row=r, end_column=LIQ_COL + 2)
    title_cell = ws.cell(row=r, column=LIQ_COL, value="LIQUIDACIÓN DE IMPUESTOS:")
    title_cell.font = Font(bold=True, size=11, color="000000")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

    # Sub-título
    r += 1
    ws.merge_cells(start_row=r, start_column=LIQ_COL, end_row=r, end_column=LIQ_COL + 2)
    sub_cell = ws.cell(row=r, column=LIQ_COL,
                         value=f"IMPUESTOS A PAGAR PERIODO {periodo_fmt[:4]}/{periodo_fmt[4:]}")
    sub_cell.fill = DARK_NAVY
    sub_cell.font = WHITE_FONT
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    sub_cell.border = _border_all()
    ws.row_dimensions[r].height = 22

    r += 1  # línea vacía
    for col in range(LIQ_COL, LIQ_COL + 3):
        ws.cell(row=r, column=col).border = _border_all()

    # Fórmulas de liquidación referenciando las filas calculadas
    end_v = max(header_v_row + 1, total_row_v - 1)
    end_c = max(r_compras_hdr + 1, total_row_c - 1)

    r += 1
    r_ventas_liq = r
    _liq_row(ws, r, LIQ_COL, "TOTAL VENTAS", "S/", f"=SUM(J{header_v_row+1}:J{end_v})",
             font_label=BOLD_FONT, font_val=BOLD_FONT)

    r += 1
    r_compras_liq = r
    _liq_row(ws, r, LIQ_COL, "TOTAL COMPRAS", "S/", f"=SUM(J{r_compras_hdr+1}:J{end_c})",
             font_label=BOLD_FONT, font_val=BOLD_FONT)

    r += 1  # separador
    for col in range(LIQ_COL, LIQ_COL + 3):
        ws.cell(row=r, column=col).border = _border_all()

    r += 1
    r_renta = r
    _liq_row(ws, r, LIQ_COL, "RENTA MYPE 1%", "S/", f"=C{r_ventas_liq}*0.01",
             font_label=BOLD_FONT, font_val=BOLD_FONT)

    r += 1
    r_igv_ventas = r
    _liq_row(ws, r, LIQ_COL, "IGV de ventas", "S/", f"=SUM(G{header_v_row+1}:G{end_v})",
             font_label=NORMAL_FONT, font_val=NORMAL_FONT)

    r += 1
    r_credito = r
    # El crédito fiscal es el IGV de Compras, es decir, el total de la columna G en la tabla COMPRAS
    _liq_row(ws, r, LIQ_COL, "IGV de compras", "-S/", f"=G{total_row_c}",
             font_label=NORMAL_FONT, font_val=NORMAL_FONT)

    r += 1
    r_igv_pagar = r
    _liq_row(ws, r, LIQ_COL, "IGV A PAGAR", "S/", f"=MAX(0, C{r_igv_ventas}-C{r_credito})",
             font_label=BOLD_FONT, font_val=BOLD_FONT,
             num_fmt='"S/"#,##0.00')

    r += 1  # separador
    for col in range(LIQ_COL, LIQ_COL + 3):
        ws.cell(row=r, column=col).border = _border_all()

    r += 1
    r_total_pagar = r
    _liq_row(ws, r, LIQ_COL, "TOTAL A PAGAR", "S/", f"=C{r_renta}+C{r_igv_pagar}",
             fill_a=YELLOW_FILL,
             font_label=Font(bold=True, size=11),
             font_val=Font(bold=True, size=11))

    r += 1
    r_margen = r
    r_utilidad = r + 1
    _liq_row(ws, r, LIQ_COL, "MARGEN DE UTILIDAD BRUTA:", "", f"=IF(C{r_ventas_liq}>0, C{r_utilidad}/C{r_ventas_liq}, 0)",
             font_label=BOLD_FONT, font_val=BOLD_FONT, num_fmt="0.00%")

    r += 1
    _liq_row(ws, r, LIQ_COL, "UTILIDAD:", "S/", f"=C{r_ventas_liq}-C{r_compras_liq}",
             font_label=BOLD_FONT, font_val=BOLD_FONT)

    # Ancho de columnas unificado para la hoja completa
    col_widths = [14, 8, 12, 14, 38, 14, 12, 12, 12, 13, 7]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Guardar y devolver ───────────────────────────────────────────────────
    filename = f"Preliminar_{req.ruc}_{req.periodo}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.get("/api/export/excel/{cliente_id}/{periodo}")
def export_excel(cliente_id: str, periodo: str):
    """Exporta las compras y ventas enriquecidas a Excel."""
    import pandas as pd
    import io
    supabase = get_supabase()
    
    # Obtener el cliente para el nombre de archivo
    res_cli = supabase.table("clientes").select("razon_social, ruc").eq("id", cliente_id).execute()
    if not res_cli.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
        
    ruc = res_cli.data[0]["ruc"]
    razon_social = res_cli.data[0]["razon_social"].replace(" ", "_")
    
    # Obtener compras
    res_compras = supabase.table("sire_preliminar_compras").select("*").eq("cliente_id", cliente_id).eq("periodo", periodo).execute()
    df_compras = pd.DataFrame(res_compras.data)
    if not df_compras.empty and "created_at" in df_compras.columns:
        df_compras = df_compras.drop(columns=["created_at", "cliente_id", "id", "error_log"], errors="ignore")
        
    # Obtener ventas
    res_ventas = supabase.table("sire_preliminar_ventas").select("*").eq("cliente_id", cliente_id).eq("periodo", periodo).execute()
    df_ventas = pd.DataFrame(res_ventas.data)
    if not df_ventas.empty and "created_at" in df_ventas.columns:
        df_ventas = df_ventas.drop(columns=["created_at", "cliente_id", "id", "error_log"], errors="ignore")
        
    # Generar Excel en memoria
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        if df_compras.empty:
            pd.DataFrame([{"Mensaje": "Sin registros"}]).to_excel(writer, sheet_name="COMPRAS", index=False)
        else:
            df_compras.to_excel(writer, sheet_name="COMPRAS", index=False)
            
        if df_ventas.empty:
            pd.DataFrame([{"Mensaje": "Sin registros"}]).to_excel(writer, sheet_name="VENTAS", index=False)
        else:
            df_ventas.to_excel(writer, sheet_name="VENTAS", index=False)
            
    buf.seek(0)
    filename = f"Preliminar_{ruc}_{periodo}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

class ManualMatchRequest(BaseModel):
    ruc: str
    periodo: str
    tipo_libro: str

@app.post("/api/sire/manual-xml-match")
def manual_xml_match(req: ManualMatchRequest):
    import sys
    from pathlib import Path
    
    script_path = Path(__file__).parent / "brain" / "db" / "sire_xml_manual_enricher.py"
    cmd = [
        sys.executable,
        str(script_path),
        "--ruc", req.ruc,
        "--periodo", req.periodo,
        "--tipo", req.tipo_libro
    ]
    
    task_id = f"manual_match_{req.ruc}_{req.periodo}_{req.tipo_libro}"
    
    # Executar de forma asíncrona usando threading como los otros bots
    thread = threading.Thread(target=_run_sync_process, args=(task_id, cmd, str(Path(__file__).parent.parent)))
    thread.daemon = True
    thread.start()
    
    return {"message": "Iniciado el acoplamiento manual de XMLs.", "task_id": task_id}


# ─────────────────────────────────────────────
# RESETEAR ENRIQUECIMIENTO XML
# ─────────────────────────────────────────────

class ResetEnriquecimientoRequest(BaseModel):
    ruc: str
    periodo: str
    tipo_libro: str | None = None  # "VENTAS", "COMPRAS", o None = ambos

@app.post("/api/enriquecimiento-xml/reset")
def reset_enriquecimiento_xml(req: ResetEnriquecimientoRequest):
    """Borra el enriquecimiento XML (estado_enriquecimiento, descripcion_comprobante, detraccion)
    de todos los comprobantes del cliente+periodo indicados,
    para poder volver a extraer la información de los XML desde cero.
    """
    supabase = get_supabase()

    # Resolver cliente_id
    res_cli = supabase.table("clientes").select("id").eq("ruc", req.ruc).execute()
    if not res_cli.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente_id = res_cli.data[0]["id"]

    campos_reset = {
        "estado_enriquecimiento": None,
        "descripcion_comprobante": None,
        "detraccion": None,
    }

    totales = {"VENTAS": 0, "COMPRAS": 0}
    tablas = []

    tipo = (req.tipo_libro or "").upper()
    if tipo == "VENTAS":
        tablas = [("sire_preliminar_ventas", "VENTAS")]
    elif tipo == "COMPRAS":
        tablas = [("sire_preliminar_compras", "COMPRAS")]
    else:
        tablas = [
            ("sire_preliminar_ventas", "VENTAS"),
            ("sire_preliminar_compras", "COMPRAS"),
        ]

    for tabla, label in tablas:
        # Solo resetear los que ya tenían estado de enriquecimiento (evitar tocar los que nunca se enriquecieron)
        res = supabase.table(tabla) \
            .update(campos_reset) \
            .eq("cliente_id", cliente_id) \
            .eq("periodo", req.periodo) \
            .not_.is_("estado_enriquecimiento", "null") \
            .execute()
        count = len(res.data) if res.data else 0
        totales[label] = count
        print(f"[reset-xml] {label}: {count} registros limpiados (cliente {req.ruc}, periodo {req.periodo})")

    return {
        "ok": True,
        "ruc": req.ruc,
        "periodo": req.periodo,
        "limpiados": totales,
        "mensaje": f"Enriquecimiento XML eliminado: {totales['VENTAS']} ventas y {totales['COMPRAS']} compras."
    }

# ─────────────────────────────────────────────
# RESETEAR CLASIFICACIÓN IA
# ─────────────────────────────────────────────

class ResetClasificacionRequest(BaseModel):
    ruc: str
    periodo: str
    tipo_libro: str | None = None  # "VENTAS", "COMPRAS", o None = ambos

@app.post("/api/clasificacion-ia/reset")
def reset_clasificacion_ia(req: ResetClasificacionRequest):
    """Borra la clasificación IA (cuenta_contable, descripcion_cuenta, categoria)
    de todos los comprobantes del cliente+periodo indicados,
    para poder volver a correr el clasificador desde cero.
    """
    supabase = get_supabase()

    # Resolver cliente_id
    res_cli = supabase.table("clientes").select("id").eq("ruc", req.ruc).execute()
    if not res_cli.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente_id = res_cli.data[0]["id"]

    campos_reset = {
        "cuenta_contable": None,
        "descripcion_cuenta": None,
        "categoria": None,
    }

    totales = {"VENTAS": 0, "COMPRAS": 0}
    tablas = []

    tipo = (req.tipo_libro or "").upper()
    if tipo == "VENTAS":
        tablas = [("sire_preliminar_ventas", "VENTAS")]
    elif tipo == "COMPRAS":
        tablas = [("sire_preliminar_compras", "COMPRAS")]
    else:
        tablas = [
            ("sire_preliminar_ventas", "VENTAS"),
            ("sire_preliminar_compras", "COMPRAS"),
        ]

    for tabla, label in tablas:
        # Solo resetear los que ya tenían cuenta asignada (evitar tocar los que nunca se clasificaron)
        res = supabase.table(tabla) \
            .update(campos_reset) \
            .eq("cliente_id", cliente_id) \
            .eq("periodo", req.periodo) \
            .not_.is_("cuenta_contable", "null") \
            .execute()
        count = len(res.data) if res.data else 0
        totales[label] = count
        print(f"[reset-ia] {label}: {count} registros limpiados (cliente {req.ruc}, periodo {req.periodo})")

    return {
        "ok": True,
        "ruc": req.ruc,
        "periodo": req.periodo,
        "limpiados": totales,
        "mensaje": f"Clasificación IA eliminada: {totales['VENTAS']} ventas y {totales['COMPRAS']} compras."
    }


# ─────────────────────────────────────────────
# RESETEAR PRELIMINAR SIRE
# ─────────────────────────────────────────────

@app.post("/api/preliminar/reset")
def reset_preliminar_sire(req: ResetClasificacionRequest):
    """Borra la data preliminar del SIRE y los registros físicos asociados
    del cliente+periodo indicados, para poder volver a subirlos/cargarlos.
    """
    supabase = get_supabase()

    # Resolver cliente_id
    res_cli = supabase.table("clientes").select("id").eq("ruc", req.ruc).execute()
    if not res_cli.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente_id = res_cli.data[0]["id"]

    tipo = (req.tipo_libro or "").upper()
    
    # 1. Primero borrar de sire_comprobantes_fisicos por dependencias
    query_fisicos = supabase.table("sire_comprobantes_fisicos").delete().eq("cliente_id", cliente_id).eq("periodo", req.periodo)
    if tipo in ("VENTAS", "COMPRAS"):
        query_fisicos = query_fisicos.eq("tipo_libro", tipo)
    query_fisicos.execute()

    totales = {"VENTAS": 0, "COMPRAS": 0}
    tablas = []

    if tipo == "VENTAS":
        tablas = [("sire_preliminar_ventas", "VENTAS")]
    elif tipo == "COMPRAS":
        tablas = [("sire_preliminar_compras", "COMPRAS")]
    else:
        tablas = [
            ("sire_preliminar_ventas", "VENTAS"),
            ("sire_preliminar_compras", "COMPRAS"),
        ]

    for tabla, label in tablas:
        res = supabase.table(tabla) \
            .delete() \
            .eq("cliente_id", cliente_id) \
            .eq("periodo", req.periodo) \
            .execute()
        count = len(res.data) if res.data else 0
        totales[label] = count
        print(f"[reset-preliminar] {label}: {count} registros eliminados (cliente {req.ruc}, periodo {req.periodo})")

    return {
        "ok": True,
        "ruc": req.ruc,
        "periodo": req.periodo,
        "limpiados": totales,
        "mensaje": f"Preliminar SIRE eliminado: {totales['VENTAS']} ventas y {totales['COMPRAS']} compras."
    }


# ─────────────────────────────────────────────
# GESTIÓN MANUAL DE COMPROBANTES FÍSICOS
# ─────────────────────────────────────────────

ESTADOS_VALIDOS = {"PENDIENTE", "DESCARGADO", "ERROR", "NO_EXISTE", "NO_DESCARGABLE", "DESFASADO"}

class UpdateEstadoRequest(BaseModel):
    estado_xml: str | None = None
    estado_pdf: str | None = None
    reset_reintentos: bool = True

@app.patch("/api/comprobante/{comprobante_id}/estado")
def update_comprobante_estado(comprobante_id: str, req: UpdateEstadoRequest):
    """Actualiza manualmente el estado XML y/o PDF de un comprobante físico.
    Permite por ejemplo pasar de NO_EXISTE a PENDIENTE para forzar un reintento.
    """
    supabase = get_supabase()

    # Validaciones
    if req.estado_xml and req.estado_xml not in ESTADOS_VALIDOS:
        raise HTTPException(status_code=400, detail=f"estado_xml inválido. Valores permitidos: {ESTADOS_VALIDOS}")
    if req.estado_pdf and req.estado_pdf not in ESTADOS_VALIDOS:
        raise HTTPException(status_code=400, detail=f"estado_pdf inválido. Valores permitidos: {ESTADOS_VALIDOS}")
    if not req.estado_xml and not req.estado_pdf:
        raise HTTPException(status_code=400, detail="Debe indicar al menos estado_xml o estado_pdf")

    # Verificar que el comprobante existe
    check = supabase.table("sire_comprobantes_fisicos").select("*").eq("id", comprobante_id).execute()
    if not check.data:
        raise HTTPException(status_code=404, detail="Comprobante no encontrado")

    comp = check.data[0]
    update_data = {}
    if req.estado_xml:
        update_data["estado_xml"] = req.estado_xml
    if req.estado_pdf:
        update_data["estado_pdf"] = req.estado_pdf
    if req.reset_reintentos:
        update_data["reintentos"] = 0
        update_data["error_log"] = None

    # Eliminar archivo físico SOLO si cambiamos el estado a PENDIENTE o ERROR.
    # NO borrar si el estado es DESCARGADO, NO_DESCARGABLE, NO_EXISTE o DESFASADO.
    ESTADOS_SIN_BORRADO = {"DESCARGADO", "NO_DESCARGABLE", "NO_EXISTE", "DESFASADO"}

    base_name = f"{comp.get('ruc_tercero', '')}-{comp.get('tipo_cp', '')}-{comp.get('serie', '')}-{comp.get('numero', '')}"
    downloads_dir = Path(__file__).parent.parent / "downloads"

    if downloads_dir.exists():
        if req.estado_xml and req.estado_xml not in ESTADOS_SIN_BORRADO:
            for f in downloads_dir.rglob(f"{base_name}.*"):
                if f.suffix.lower() in ('.xml', '.zip'):
                    try:
                        f.unlink()
                    except Exception:
                        pass
        if req.estado_pdf and req.estado_pdf not in ESTADOS_SIN_BORRADO:
            for f in downloads_dir.rglob(f"{base_name}.*"):
                if f.suffix.lower() == '.pdf':
                    try:
                        f.unlink()
                    except Exception:
                        pass

    supabase.table("sire_comprobantes_fisicos").update(update_data).eq("id", comprobante_id).execute()
    return {"ok": True, "id": comprobante_id, "updated": update_data}


# ─────────────────────────────────────────────
# RESET MASIVO A PENDIENTE (excluyendo DESCARGADO)
# ─────────────────────────────────────────────

class ResetPendientesRequest(BaseModel):
    ruc: str
    periodo: str | None = None
    tipo_libro: str | None = None   # "VENTAS", "COMPRAS" o None = ambos

@app.post("/api/comprobantes/reset-pendientes")
def reset_comprobantes_pendientes(req: ResetPendientesRequest):
    """Pone en PENDIENTE todos los comprobantes físicos cuyo estado_xml o estado_pdf
    NO sea DESCARGADO. Excluye los comprobantes ya descargados.
    Útil para forzar un nuevo intento masivo del bot descargador.
    """
    supabase = get_supabase()

    # Resolver cliente_id
    res_cli = supabase.table("clientes").select("id").eq("ruc", req.ruc).execute()
    if not res_cli.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    cliente_id = res_cli.data[0]["id"]

    # Construir la query base para estados != DESCARGADO y != DESFASADO
    # Para XML: resetear si NO está DESCARGADO ni DESFASADO
    # Para PDF: resetear si NO está DESCARGADO ni DESFASADO
    query_xml = (
        supabase.table("sire_comprobantes_fisicos")
        .update({"estado_xml": "PENDIENTE", "reintentos": 0, "error_log": None})
        .eq("cliente_id", cliente_id)
        .neq("estado_xml", "DESCARGADO")
        .neq("estado_xml", "DESFASADO")
    )
    query_pdf = (
        supabase.table("sire_comprobantes_fisicos")
        .update({"estado_pdf": "PENDIENTE", "reintentos": 0, "error_log": None})
        .eq("cliente_id", cliente_id)
        .neq("estado_pdf", "DESCARGADO")
        .neq("estado_pdf", "DESFASADO")
    )

    if req.periodo:
        query_xml = query_xml.eq("periodo", req.periodo)
        query_pdf = query_pdf.eq("periodo", req.periodo)

    if req.tipo_libro:
        query_xml = query_xml.eq("tipo_libro", req.tipo_libro.upper())
        query_pdf = query_pdf.eq("tipo_libro", req.tipo_libro.upper())

    res_xml = query_xml.execute()
    res_pdf = query_pdf.execute()

    total_xml = len(res_xml.data) if res_xml.data else 0
    total_pdf = len(res_pdf.data) if res_pdf.data else 0

    print(f"[reset-pendientes] {req.ruc} | periodo={req.periodo} | tipo={req.tipo_libro}")
    print(f"  → XML reseteados: {total_xml} | PDF reseteados: {total_pdf}")

    return {
        "ok": True,
        "ruc": req.ruc,
        "periodo": req.periodo,
        "tipo_libro": req.tipo_libro,
        "reseteados_xml": total_xml,
        "reseteados_pdf": total_pdf,
        "mensaje": f"Se pusieron en PENDIENTE {total_xml} XML y {total_pdf} PDF (excluyendo DESCARGADO)."
    }
@app.post("/api/comprobante/{comprobante_id}/upload")
async def upload_comprobante_file(
    comprobante_id: str,
    file: UploadFile = File(...),
    file_type: str = Form(...),  # "pdf" o "xml"
):
    """Sube manualmente un archivo PDF o XML para un comprobante físico.
    Guarda el archivo en la carpeta de downloads del cliente y actualiza la ruta en DB.
    """
    supabase = get_supabase()

    # Verificar que el comprobante existe y obtener sus datos
    check = supabase.table("sire_comprobantes_fisicos") \
        .select("id, cliente_id, periodo, tipo_libro, serie, numero, tipo_cp, ruc_tercero, clientes!inner(ruc, razon_social)") \
        .eq("id", comprobante_id) \
        .execute()

    if not check.data:
        raise HTTPException(status_code=404, detail="Comprobante no encontrado")

    comp = check.data[0]
    cliente = comp.get("clientes", {})
    ruc_cliente = cliente.get("ruc", "unknown")
    rs_cliente = (cliente.get("razon_social") or "").strip().replace("/", "-").replace("\\", "-")
    
    # Sanitizar nombre de carpeta (igual que en el scraper)
    import re
    rs_safe = re.sub(r'[\\/*?"<>|]', "", rs_cliente).strip()
    folder_client = f"{rs_safe} {ruc_cliente}".strip()

    # Validar tipo
    file_type = file_type.lower()
    if file_type not in ("pdf", "xml"):
        raise HTTPException(status_code=400, detail="file_type debe ser 'pdf' o 'xml'")

    # Determinar extensión real
    original_name = file.filename or ""
    ext = Path(original_name).suffix.lower() or f".{file_type}"
    if ext not in (".pdf", ".xml", ".zip"):
        ext = f".{file_type}"

    # Construir ruta destino (misma estructura que el scraper)
    root_dir = Path(__file__).parent.parent
    period = comp.get("periodo", "unknown")
    book = "sales" if comp.get("tipo_libro") == "VENTAS" else "purchases"
    subfolder = "pdf" if file_type == "pdf" else "xml"
    
    ruc_tercero = (comp.get("ruc_tercero") or "").strip()
    if ruc_tercero == "-":
        ruc_tercero = ""
    ruc_tercero = ruc_tercero or ruc_cliente
    
    tipo_cp = comp.get("tipo_cp", "00")
    serie = comp.get("serie", "")
    numero = comp.get("numero", "")
    base_name = f"{ruc_tercero}-{tipo_cp}-{serie}-{numero}"

    dest_dir = root_dir / "downloads" / "xml" / folder_client / period / book / subfolder
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_path = dest_dir / f"{base_name}{ext}"

    # Escribir el archivo
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    with open(dest_path, "wb") as f:
        f.write(contents)

    # Actualizar DB
    update_data: dict = {}
    if file_type == "pdf":
        update_data["ruta_pdf"] = str(dest_path)
        update_data["estado_pdf"] = "DESCARGADO"
    else:
        update_data["ruta_xml"] = str(dest_path)
        update_data["estado_xml"] = "DESCARGADO"
    update_data["reintentos"] = 0
    update_data["error_log"] = None

    supabase.table("sire_comprobantes_fisicos").update(update_data).eq("id", comprobante_id).execute()

    return {
        "ok": True,
        "id": comprobante_id,
        "file_type": file_type,
        "saved_to": str(dest_path),
        "size_bytes": len(contents),
    }


# ─────────────────────────────────────────────


# ─────────────────────────────────────────────
# EMISIÓN DE FACTURAS - INTEGRACIÓN REAL APISUNAT
# ─────────────────────────────────────────────
import httpx

APISUNAT_SANDBOX = "https://sandbox.apisunat.pe"
APISUNAT_PROD    = "https://app.apisunat.pe"

class FacturacionEmitirRequest(BaseModel):
    emisor_ruc: str
    receptor: dict
    comprobante: dict
    items: list
    totales: dict
    token: str         # Bearer token del cliente - obtenido en app.apisunat.pe
    sandbox: bool = True  # True = pruebas, False = producción

@app.post("/api/facturacion/emitir")
def emitir_comprobante(req: FacturacionEmitirRequest):
    if not req.token:
        raise HTTPException(status_code=400, detail="Se requiere el token de APISUNAT de esta empresa.")

    # Construir el payload que acepta APISUNAT
    # Mapear tipo de comprobante
    tipo_doc_map = {"01": "factura", "03": "boleta"}
    tipo_doc = tipo_doc_map.get(req.comprobante.get("tipo", "01"), "factura")

    # Construir items en formato APISUNAT
    apisunat_items = []
    for item in req.items:
        cantidad = float(item.get("cantidad", 1))
        precio_unitario = float(item.get("precio_unitario", 0))
        # valor_unitario = precio sin IGV
        valor_unitario = precio_unitario / 1.18
        apisunat_items.append({
            "unidad_de_medida": "NIU",
            "descripcion": item.get("descripcion", "Producto"),
            "cantidad": str(cantidad),
            "valor_unitario": f"{valor_unitario:.6f}",
            "porcentaje_igv": "18",
            "codigo_tipo_afectacion_igv": "10",
            "nombre_tributo": "IGV"
        })

    payload = {
        "documento": tipo_doc,
        "serie": req.comprobante.get("serie", "F001"),
        "numero": int(req.comprobante.get("correlativo", 1) or 1),
        "fecha_de_emision": req.comprobante.get("fecha"),
        "moneda": req.comprobante.get("moneda", "PEN"),
        "tipo_operacion": "0101",
        "cliente_tipo_de_documento": "6" if len(req.receptor.get("ruc","")) == 11 else "1",
        "cliente_numero_de_documento": req.receptor.get("ruc", ""),
        "cliente_denominacion": req.receptor.get("razon_social", ""),
        "cliente_direccion": req.receptor.get("direccion", ""),
        "items": apisunat_items,
        "total": f"{req.totales.get('total', 0):.2f}"
    }

    base_url = APISUNAT_SANDBOX if req.sandbox else APISUNAT_PROD
    url = f"{base_url}/api/v3/documents"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {req.token}"
    }

    try:
        response = httpx.post(url, json=payload, headers=headers, timeout=30)
        data = response.json()

        if response.status_code == 200 and data.get("success"):
            payload_resp = data.get("payload", {})
            return {
                "status": "success",
                "estado": payload_resp.get("estado"),
                "message": data.get("message"),
                "xml_url": payload_resp.get("xml"),
                "pdf_url": payload_resp.get("pdf"),
                "cdr_url": payload_resp.get("cdr"),
                "hash": payload_resp.get("hash"),
            }
        else:
            # Devolver el error tal como viene de APISUNAT
            raise HTTPException(
                status_code=response.status_code,
                detail=data.get("message", "Error al emitir en APISUNAT")
            )
    except httpx.TimeoutException:
        raise HTTPException(status_code=504, detail="Tiempo de espera agotado conectando con APISUNAT.")
    except httpx.RequestError as e:
        raise HTTPException(status_code=502, detail=f"Error de conexión con APISUNAT: {str(e)}")

@app.post("/api/facturacion/estado")
def consultar_estado(body: dict):
    """Consulta el estado de un comprobante ya emitido en APISUNAT"""
    token = body.get("token", "")
    sandbox = body.get("sandbox", True)
    if not token:
        raise HTTPException(status_code=400, detail="Token requerido")

    base_url = APISUNAT_SANDBOX if sandbox else APISUNAT_PROD
    url = f"{base_url}/api/v3/status"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}"
    }
    payload = {
        "serie": body.get("serie"),
        "numero": body.get("numero"),
        "tipo": body.get("tipo", "01"),
        "ruc_emisor": body.get("ruc_emisor")
    }
    try:
        response = httpx.post(url, json=payload, headers=headers, timeout=20)
        return response.json()
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))
